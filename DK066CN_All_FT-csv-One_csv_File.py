# -*- coding: utf-8 -*-
"""
Created on Tue Aug 19 09:13:38 2025

@author: pc
"""

""" # GPT中文产生
import os
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def process_cn_files(folder_path, output_csv, output_xlsx):
    target_cols = ['IG_12V', 'IG_20V', 'IG_22V', 'VG_TH', 'ID_BV', 'Rdson']
    all_data = []

    # 遍历文件夹中所有包含"CN"的csv文件
    for filename in os.listdir(folder_path):
        if filename.endswith('.csv') and 'CN' in filename:
            file_path = os.path.join(folder_path, filename)
            print(f"处理文件: {filename}")

            # 先读取文件所有行，定位表头行（包含SITE_NUM）
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            header_line_idx = None
            for i, line in enumerate(lines):
                if 'SITE_NUM' in line:
                    header_line_idx = i
                    break

            if header_line_idx is None:
                print(f"文件 {filename} 未找到包含 SITE_NUM 的表头行，跳过")
                continue

            # 正式数据从 header_line_idx + 5 行开始读取
            data_start = header_line_idx + 5

            # 读取数据，header=header_line_idx，跳过前data_start行
            try:
                df = pd.read_csv(file_path, header=header_line_idx, skiprows=range(header_line_idx+1, data_start))
            except Exception as e:
                print(f"读取文件 {filename} 出错: {e}")
                continue

            # 过滤需要的列，忽略缺失列
            cols_exist = [col for col in target_cols if col in df.columns]
            if not cols_exist:
                print(f"文件 {filename} 不包含目标列，跳过")
                continue

            df_filtered = df[cols_exist].copy()
            df_filtered['source_file'] = filename  # 标记来源文件
            all_data.append(df_filtered)

    if not all_data:
        print("没有找到符合条件的数据")
        return

    # 合并所有数据
    merged_df = pd.concat(all_data, ignore_index=True)
    merged_df.to_csv(output_csv, index=False, encoding='utf-8-sig')
    print(f"合并数据保存到 {output_csv}")

    # 创建Excel文件保存直方图
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    for col in target_cols:
        if col not in merged_df.columns:
            print(f"列 {col} 不存在，跳过绘图")
            continue

        data = merged_df[col].dropna()

        plt.figure(figsize=(6,4))
        plt.hist(data, bins=30, density=True, alpha=0.6, color='g')
        plt.title(f'Histogram of {col}')
        plt.xlabel(col)
        plt.ylabel('Density')

        # 保存图像到内存
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png')
        plt.close()
        img_buffer.seek(0)

        # 新建sheet，插入图片
        ws = wb.create_sheet(title=col)
        img = Image(img_buffer)
        img.anchor = 'A1'
        ws.add_image(img)

    wb.save(output_xlsx)
    print(f"直方图保存到 {output_xlsx}")

# 使用示例
folder_path = r'c:\MiddleTest\DK066\FT\CN'  # 替换为你的文件夹路径
output_csv = r'FT_CN_total.csv'
output_xlsx = r'FT_CN_histograms.xlsx'

process_cn_files(folder_path, output_csv, output_xlsx)
"""


# deepseek产生
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import re
import warnings

#提交GitHub标记2

# 配置参数
folder_path = r'c:\MiddleTest\DK066\FT\CN'  # 替换为你的文件夹路径
output_csv = folder_path+r'\FT_CN_total.csv'

output_excel = folder_path+r'\Histograms.xlsx'
target_columns = ['IG_12V', 'IG_20V', 'IG_22V', 'VG_TH', 'ID_BV', 'Rdson']
required_keyword = 'SITE_NUM'
skip_after_header = 5

file_pattern1 = 'CN'
file_pattern = 'ISG6133'
output_csv = folder_path+rf'\FT_{file_pattern}_{file_pattern1}_total.csv'
output_excel = folder_path+rf'\Histograms_{file_pattern}_{file_pattern1}.xlsx'

# 支持的编码格式列表（按优先级排序）
ENCODINGS = ['utf-8', 'latin1', 'iso-8859-1', 'cp1252', 'gbk', 'gb18030', 'big5']

# 存储所有数据
all_data = []

def detect_encoding(filepath):
    """尝试检测文件的正确编码"""
    for encoding in ENCODINGS:
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                f.read(1024)  # 读取前1024字节测试
            return encoding
        except:
            continue
    return 'utf-8'  # 默认回退到utf-8

# 遍历文件夹中的CSV文件
for filename in os.listdir(folder_path):
    if filename.endswith('.csv') and file_pattern in filename and file_pattern1 in filename:
        filepath = os.path.join(folder_path, filename)
        print(f"\n处理文件: {filename}")
        
        # 检测文件编码
        file_encoding = detect_encoding(filepath)
        print(f"检测到编码: {file_encoding}")
        
        # 查找表头行
        header_row = None
        try:
            with open(filepath, 'r', encoding=file_encoding, errors='replace') as f:
                for i, line in enumerate(f):
                    if required_keyword in line:
                        header_row = i
                        print(f"找到表头行: {header_row}")
                        break
        except Exception as e:
            print(f"读取文件头错误: {str(e)}")
            continue
        
        if header_row is None:
            print(f"警告: 未找到表头关键字 '{required_keyword}'，跳过")
            continue
        
        # 计算数据起始行
        data_start = header_row + skip_after_header + 1
        
        try:
            # 读取CSV文件 - 更兼容的格式处理
            df = pd.read_csv(
                filepath, 
                #header=header_row,
                skiprows=lambda x: x < header_row or x in range(header_row+1, header_row+4),
                encoding=file_encoding,
                engine='python',  # 使用Python引擎处理格式问题
                dtype=str,  # 先作为字符串读取
                na_filter=False,  # 不过滤空值
                on_bad_lines='skip'  # 警告格式错误的行
            )
            
            # 将数值列转换为float
            for col in target_columns:
                if col in df.columns:
                    # 尝试转换为数值类型
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # 检查目标列是否存在
            missing_cols = [col for col in target_columns if col not in df.columns]
            if missing_cols:
                print(f"警告: 缺少列 {missing_cols}，跳过")
                continue
            
            # 提取目标列
            extracted = df[target_columns].copy().dropna(how='all')
            #extracted['Source_File'] = filename  # 添加来源标记
            all_data.append(extracted)
            print(f"提取成功: 找到 {len(extracted)} 行数据")
            
        except Exception as e:
            print(f"处理文件时出错: {str(e)}")

# 合并所有数据
if all_data:
    combined_df = pd.concat(all_data, ignore_index=True)
    
    # 保存合并数据前进行数据清洗
    for col in target_columns:
        if col in combined_df.columns:
            # 移除无效值
            combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce')
            # 删除完全无效的列
            if combined_df[col].isna().all():
                print(f"警告: 列 {col} 全部为无效值，将被移除")
                combined_df.drop(col, axis=1, inplace=True)
    
    # 保存合并数据
    combined_df.to_csv(output_csv, index=False)
    print(f"\n合并数据已保存至: {output_csv} (总行数: {len(combined_df)})")
    
    # 创建Excel工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认工作表
    
    # 为每列创建直方图
    for col in target_columns:
        if col not in combined_df.columns:
            print(f"警告: 列 {col} 不存在于合并数据中，跳过直方图生成")
            continue
            
        # 清理列名用于工作表名称
        clean_name = re.sub(r'[\\/*?:[\]]', '', col)[:30]
        
        # 创建新工作表
        ws = wb.create_sheet(title=clean_name)
        
        # 生成直方图
        plt.figure(figsize=(10, 6))
        
        # 提取有效数据
        if col == 'IG_12V':
            valid_data = combined_df[col].dropna()
            valid_data = valid_data[valid_data>0.1][valid_data<0.8]
        elif col == 'IG_20V':
            valid_data = combined_df[col].dropna()
            valid_data = valid_data[valid_data>0.2][valid_data<1.1]
        elif col == 'IG_22V':
            valid_data = combined_df[col].dropna()
            valid_data = valid_data[valid_data>0.3][valid_data<1.2]
        elif col == 'VG_TH':
            valid_data = combined_df[col].dropna()
            valid_data = valid_data[valid_data>3.1][valid_data<4.5]
        elif col == 'ID_BV':
            valid_data = combined_df[col].dropna()
            valid_data = valid_data[valid_data>=0][valid_data<20]
        elif col == 'Rdson':
            valid_data = combined_df[col].dropna()
            valid_data = valid_data[valid_data>=90][valid_data<500]
        
        if len(valid_data) > 0:
            plt.hist(valid_data, bins=30, density=True, 
                     alpha=0.6, color='g', edgecolor='black')
            plt.title(f'Distribution of {col}')
            plt.xlabel(col)
            plt.ylabel('Density')
            plt.grid(True, alpha=0.3)
        else:
            # 如果没有有效数据，创建空白图表
            plt.text(0.5, 0.5, f'No valid data for {col}', 
                     ha='center', va='center', fontsize=12)
            plt.title(f'No Data: {col}')
        
        # 保存图像到BytesIO
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')
        plt.close()
        
        # 将图像插入Excel
        img_buffer.seek(0)
        img = Image(img_buffer)
        img.width = 600
        img.height = 400
        ws.add_image(img, 'A1')
        ws.column_dimensions['A'].width = 20  # 调整列宽
    
    # 保存Excel文件
    wb.save(output_excel)
    print(f"直方图已保存至: {output_excel}")

else:
    print("未找到有效数据，请检查文件路径和过滤条件")
