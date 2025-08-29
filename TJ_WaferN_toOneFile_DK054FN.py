# -*- coding: utf-8 -*-
"""
Created on Thu Jun 26 13:05:15 2025

@author: deepseek
"""

import os
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
# from openpyxl.styles import PatternFill, Border, Side, Color

# 拷贝*.xlsm中Ivcc工作表（sheet），建立n个参数的底板文件
from openpyxl import load_workbook


# 配置路径
# path = r"c:\MiddleTest\DK054\FN版\DK05400FN_AS29R7.1"
# path = r"c:\MiddleTest\DK054\FN版\DK05400FN_AS29R6.1"
# path = r"c:\MiddleTest\DK054\D版\AR6SRN.1"
# path = r"c:\MiddleTest\DK054\D版\AR5ARW.1"
path = r"c:\MiddleTest\DK054\FN测成D版"
pathnum = r"AS4HR9.1"

#os.makedirs(path, exist_ok=True)
ProgramName_testTime = "测试程序名-时间-Total-Pass-Fail.csv"
param_MaxMin_Unit = "total参数名-Max-Min-Unit.csv"
all_Pcs_TJparam_detail= "total-5千行一片-纯详细数据-float-int64.csv"

#待分析的数据名列表
pram_columns = ['Bin', 'AA1', 'Ivcc', 'Ivcch', 'Vdd', 'Rfb', 'Ivcch_ton', 'Vgtl', 'Vgth', 
                'Ivcch_toff', 'Irocc', 'Ivs', 'Vovp', 'Vccovp', 'Tclk',	'Trunf2',
                'Vrsth', 'Vuvlo', 'IVCC_BV',	'IVCCH_BV',	'TONMAX']

name_1 = ('Seq_Name', 'FUN', 'ISD.10', 'ISD.11', 'ISD.12', 'ISD.13', 'ISD.14', 'ISD.15', 'ISD.16', 'ISD.17', 'ISD.18', 'ISD.19', 'ISD.20', 'ISD.21', 'ISD.22', 'ISD.23', 'ISD.24', 'ISD.25')
name_2 = ('Bin','AA1', 'Vgth', 'Ivcc', 'Ivcch', 'Vdd', 'Rfb', 'Ivcch_ton', 'Vgtl', 'Ivcch_toff', 'Irocc', 'Ivs', 'Vovp', 'Vccovp', 'Tclk', 'Trunf2', 'Vrsth', 'Vuvlo')

columnsRename = dict(zip(name_1, name_2))

df1 = pd.read_csv(
    os.path.join(path, r"total参数名-Max-Min-Unit.csv"),
    # header=None,
    skiprows=lambda x: x in [1,2,3,4],
    nrows=4,
    #usecols=list(range(0,3)),
    encoding='utf-8',   #'gbk'
    #on_bad_lines='skip',
    #quotechar='"',
    #low_memory=False
)
del df1['waferNum']
columnsRename = dict(zip(list(df1.iloc[0])[:18], list(df1.columns)[:18]))

# unitList = ['Bin', 'uA', 'uA', 'V', 'Kohm', 'mA', 'mA', 'V', 'mA', 'mA', 'uA', 
#             'V', 'V', 'V', 'us', 'us', 'V', 'uA', 'uA', 'us']


# 获取所有CSV文件
all_files = [f for f in os.listdir(path) if f.endswith('.csv')]
all_files = [f for f in all_files if 'total' not in f.lower()]
all_files.sort(key=lambda s: int(re.search(rf'{pathnum[-8:]}-(\d+)',s).group(1)))
file_pcs_list = all_files #[re.search(r'CS43CX.1-(\d+)', s).group(1)+'#' for s in all_files]
numPcs = [int(re.search(rf'{pathnum[-8:]}-(\d+)',s).group(1)) for s in all_files]

# 识别样本ID
sample_ids = set()
for file in all_files:
    matchlist = re.search(rf'{pathnum[-8:]}-(\d+)', file)
    #print(match)
    if matchlist:
        sample_ids.add(int(matchlist.group(1)))
n = len(sample_ids)
print(f"识别到样本数量: {n}个，分别为：{file_pcs_list}")


# 测试程序名、时间，Total、Pass、Fail
'''dfList = []
df = pd.read_csv(
    os.path.join(path, all_files[0]),
    header=None,
    skiprows=lambda x: x >= 4 and x<=6,  # csv行数-1
    nrows=5,
    encoding='utf-8',   #'gbk'
    #on_bad_lines='skip',
    #quotechar='"',
    #low_memory=False
)
df = pd.concat([df, pd.DataFrame([np.nan, np.nan])])
dfList.append(df)
for index,filex in enumerate(all_files, start=1):
    PCS_num = pd.DataFrame([[f'{index}#']*len(df.columns)], columns=df.columns)
    df = pd.read_csv(
        os.path.join(path, filex),
        header=None,
        skiprows=lambda x: x <= 5 or x==7,  # csv行数-1
        nrows=4,
        encoding='utf-8',   #'gbk'
        #on_bad_lines='skip',
        #quotechar='"',
        #low_memory=False
    )    
    total = int(df[0][1].split(':')[-1].strip())
    pass_count = int(df[0][2].split(':')[-1].strip())
    
    # 计算通过率
    pass_rate = pass_count / total
    formatted_pass_rate = f"{pass_rate * 100:.1f}%"
    
    # 构建新DataFrame并保持冒号对齐
    new_data = {
        0: [f'Passing rate :    {formatted_pass_rate}']
    }    
    
    empty_row = pd.DataFrame([[np.nan]*len(df.columns)], columns=df.columns)
    df = pd.concat([PCS_num, df, pd.DataFrame(new_data), empty_row])
    df.reset_index(inplace=True, drop=True)

    dfList.append(df)
    
df0 = pd.concat(dfList)
df0.reset_index(inplace=True, drop=True)
df0.to_csv( os.path.join(path, ProgramName_testTime), index=False, header=False)'''

# 参数名，Max、Min，Unit
# 方法1：从原测试数据中读取
'''df1 = pd.read_csv(
    os.path.join(path, all_files[0]),
    #header=None,
    skiprows=lambda x: x <= 12 or x == 14,  # csv行数-1
    nrows=4,
    encoding='utf-8',   #'gbk'
    #on_bad_lines='skip',
    #quotechar='"',
    #low_memory=False
)
columns_raw = df1.columns 
#更改部分列名
df1.rename(columns=columnsRename, inplace=True)
 
# 直接选择
df1 = df1[pram_columns]
df1['waferNum'] = 'wafer-#'
# 放到前面
df1 = pd.concat([df1.iloc[:, -1], df1.iloc[:, :-1]], axis=1)
df1.to_csv( os.path.join(path, param_MaxMin_Unit), index=False)'''

# 方法2：手动构建


'''df1 = pd.read_csv(
    os.path.join(path, all_files[0]),
    # header=None,
    skiprows=lambda x: x <= 12, #or x == 14,  # csv行数-1
    nrows=4,
    #usecols=list(range(0,3)),
    encoding='utf-8',   #'gbk'
    #on_bad_lines='skip',
    #quotechar='"',
    #low_memory=False
)
columns_raw = df1.columns '''

# 纯详细数据，类型统一“float”或“int64”
for index, sample_id in enumerate(all_files):
    df = pd.read_csv(
        os.path.join(path, sample_id),
        #header=None,
        skiprows=lambda x: x <= 12 or x in [14,15,16,17],  # csv行数-1
        #usecols=csv_read_columns,
        #nrows=4,
        encoding='utf-8',   #'gbk'
        #on_bad_lines='skip',
        #quotechar='"',
        low_memory=False
    )
    
    #加入列名
    #df.columns = columns_raw  
    #更改部分列名
    df.rename(columns=columnsRename, inplace=True)    
    
     
    # 直接选择
    df = df[pram_columns]
    #或丢掉
    # columns_to_drop = [col for col in df.columns if col not in pram_columns]
    # df.drop(columns=columns_to_drop, inplace=True)
    
    # 获取某一列的真实类型分布，有几种类型
    #type_counts = df['IVS'].map(type).value_counts()
    for col in pram_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
        if col in ['AA', 'FCS']:
            '''df[col] = df[col].astype(float).astype(int).astype(str).apply(lambda x: int(x, 2)).astype('Int64')'''
            df[col] = df[col].astype('Int64')
        #print(col,df[col].dtype)
    #type_counts = df['IVS'].map(type).value_counts()
    #df['IVS'].isna().sum()
    
    # 去掉NaN行
    #df.dropna(inplace=True,thresh = 30) #how = 'any',
        
    
    # 假设df是原始DataFrame（19000行）
    new_index = range(5000)  # 生成0-4999的索引
    df = df.reindex(new_index)#.fillna('')  # 扩展并填充空字符串
    
    # 增加晶圆#作为一列，好几片数据放到一块时，便于区分
    df['waferNum'] = f'{numPcs[index]}#'        
    # 放到前面
    df = pd.concat([df.iloc[:, -1], df.iloc[:, :-1]], axis=1)
    #df=df[list(df.columns[-1:])+list(df.columns[:-1])]
        
    
    #保存数据
    # 首次写入创建文件，后续追加模式
    mode = 'w' if index==0 else 'a'
    header = (index==0)  # 仅第一次写入列名
    
    # 保存到CSV（注意index=False避免重复索引）
    # df.to_csv(f"{base_filename}_{i+1}.csv", 
    #          mode=mode,
    #          header=header,
    #          index=False)
    df.to_csv( os.path.join(path, all_Pcs_TJparam_detail), mode=mode,
              index=False, header=header)
    print(f"保存完了{file_pcs_list[index]}")
        

print(f"处理完成! 结果保存至: {os.path.join(path, all_Pcs_TJparam_detail)}")
