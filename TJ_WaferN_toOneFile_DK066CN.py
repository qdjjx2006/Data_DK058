# -*- coding: utf-8 -*-
"""
Created on Thu Jun 26 13:05:15 2025

@author: deepseek
"""

import os
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
# from openpyxl.styles import PatternFill, Border, Side, Color

# 拷贝*.xlsm中Ivcc工作表（sheet），建立n个参数的底板文件
from openpyxl import load_workbook


# 配置路径
# path = r"c:\MiddleTest\DK054\FN版\DK05400FN_AS29R7.1"
# path = r"c:\MiddleTest\DK054\FN版\DK05400FN_AS29R6.1"
# path = r"c:\MiddleTest\DK054\D版\AR6SRN.1"
# path = r"c:\MiddleTest\DK054\D版\AR5ARW.1"
path = r"c:\MiddleTest\DK066\CN版\PR8G0E.8"
path = r"c:\MiddleTest\DK066\CN版\PRCP91.5"

#os.makedirs(path, exist_ok=True)
ProgramName_testTime = "测试程序名-时间-Total-Pass-Fail.csv"
param_MaxMin_Unit = "total参数名-Max-Min-Unit.csv"
all_Pcs_TJparam_detail= "total-5千行一片-纯详细数据-float-int64.csv"

#待分析的数据名列表
pram_columns = ["SOFT_BIN",	"TonMax", "IG_12V",	"VGT_12V_100UA", "VGT_12V_1MA",	"IG_20V",	"VGT_20V_100UA",	"VGT_20V_1MA",	"IG_22V",	"VGT_22V_100UA",	"VDD_D40V",	"ID_40V",	"ID_750V",	"IGT_0V",	"VDD_G12V",	"VD_SAT",	"TonD",	"ToffD",	"TonMIN",	"VGth_GTON"]


# 获取所有CSV文件
all_files = [f for f in os.listdir(path) if f.endswith('.csv')]
all_files = [f for f in all_files if 'total' not in f.lower()]
all_files = [f for f in all_files if '#' not in f.lower()]
def sort_key(filename):
    # 提取主数字（开头数字）
    main_num_match = re.match(r'^(\d+)', filename)
    main_num = int(main_num_match.group(1)) if main_num_match else float('inf')

    # 提取后缀（第一个 '-' 到 '.' 之间的部分）
    suffix_match = re.search(r'-(.+?)\.', filename)
    suffix = suffix_match.group(1) if suffix_match else ''

    return (main_num, suffix)
file_pcs_list = sorted(all_files, key=sort_key)
all_files = file_pcs_list

numPcs = [int(re.search(r'^(\d+)',s).group(1)) for s in all_files]

# 识别样本ID
sample_ids = set()
for file in all_files:
    matchlist = re.search(r'^(\d+)', file)
    #print(match)
    if matchlist:
        sample_ids.add(int(matchlist.group(1)))
n = len(sample_ids)
print(f"识别到样本数量: {n}个，分别为：{file_pcs_list}")
sample_ids_GenCsvFile = [] #  sample_ids #  

# 将每一晶圆的多次测试数据，综合到一个文件
for sample_id in sample_ids_GenCsvFile:
    print(f"\n处理样本 #{sample_id}...")
    
    # 识别文件
    prefix = str(sample_id)
    initial_files = [f for f in all_files if f.startswith(prefix) and '#' not in f.lower()
                    and 'fc' not in f.lower() and not re.match(rf'^{prefix}\d', f)]
    retest_files = [f for f in all_files if f.startswith(prefix) and 'fc' in f.lower() 
                    and not re.match(rf'^{prefix}\d', f)]
    retest_files.sort(reverse=True)
    
    # 1. 初测综合
    initial_dfs = []
    for file in initial_files:
        df = pd.read_csv(
            os.path.join(path, file),
            #header=56,
            skiprows=lambda x: x < 51 or 52 <= x <= 55,  # csv行数-1
            #skiprows = 55,
            encoding='gbk'
        )
        initial_dfs.append(df)
    
    if initial_dfs:
        initial_combined = pd.concat(initial_dfs, ignore_index=True)
        #initial_combined.to_csv(os.path.join(path, f"{sample_id}#-初测综合.csv"), index=False)
        
        # 2. 初测BIN1
        initial_bin1 = initial_combined[initial_combined["SOFT_BIN"] == 1]
        #initial_bin1.to_csv(os.path.join(path, f"{sample_id}#-初测综合bin1.csv"), index=False)
    else:
        print(f"警告: 样本 #{sample_id}# 无初测文件")
        continue

    # 3. 复测处理
    retest_bin1_dfs = []
    for file in retest_files:
        df = pd.read_csv(
            os.path.join(path, file),
            skiprows=lambda x: x < 51 or 52 <= x <= 55,
            encoding='gbk'
        )
        if file != retest_files[-1]:
            retest_bin1 = df[df["SOFT_BIN"] == 1]
        else:
            retest_bin1 = df    # 最后一次复测文件，不去掉非bin1数据
        retest_bin1_dfs.append(retest_bin1)
        
    if retest_bin1_dfs:
        retest_combined = pd.concat(retest_bin1_dfs, ignore_index=True)
        # retest_combined.to_csv(os.path.join(path, f"{sample_id}#-复测综合.csv"), index=False)
        # 4. 总综合
        total_combined = pd.concat([initial_bin1, retest_combined], ignore_index=True)
    else:
        print(f"警告: 样本 #{sample_id} 无复测文件")
        #retest_combined = pd.DataFrame(columns=initial_bin1.columns)
        # 4. 总综合
        total_combined = initial_combined

    all_bin1 = pd.concat([initial_bin1, 
        retest_combined[retest_combined["SOFT_BIN"] == 1]], ignore_index=True)
    all_bin1 = all_bin1[pram_columns]
    all_bin1.to_csv(os.path.join(path, f"{sample_id}#-all综合bin1.csv"), index=False)
    #len_list.append((len(initial_combined),len(total_combined)))
    # total_combined.to_csv(os.path.join(path, f"{sample_id}#-total综合.csv"), index=False)

# 纯详细数据，类型统一“float”或“int64”
all_files = [f for f in os.listdir(path) if f.endswith('.csv')]
all_files = [f for f in all_files if '#' in f.lower()]
for index, sample_id in enumerate(all_files):
    df = pd.read_csv(
        os.path.join(path, sample_id),
        #header=None,
        #skiprows=lambda x: x < 51 or x in [52,53,54,55],  # csv行数-1
        #usecols=csv_read_columns,
        #nrows=4,
        encoding='gbk',  #'utf-8',   #
        #on_bad_lines='skip',
        #quotechar='"',
        low_memory=False
    )
    
    #加入列名
    #df.columns = columns_raw  
    #更改部分列名
    #df.rename(columns=columnsRename, inplace=True)    
    
     
    # 直接选择
    df = df[pram_columns]
    #或丢掉
    # columns_to_drop = [col for col in df.columns if col not in pram_columns]
    # df.drop(columns=columns_to_drop, inplace=True)
    
    # 获取某一列的真实类型分布，有几种类型
    #type_counts = df['IVS'].map(type).value_counts()
    """for col in pram_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
        if col in ['AA', 'FCS']:
            '''df[col] = df[col].astype(float).astype(int).astype(str).apply(lambda x: int(x, 2)).astype('Int64')'''
            df[col] = df[col].astype('Int64')"""
        #print(col,df[col].dtype)
    #type_counts = df['IVS'].map(type).value_counts()
    #df['IVS'].isna().sum()
    
    # 去掉NaN行
    #df.dropna(inplace=True,thresh = 30) #how = 'any',
        
    
    # 假设df是原始DataFrame（19000行）
    new_index = range(70000)  # 生成0-4999的索引
    df = df.reindex(new_index)#.fillna('')  # 扩展并填充空字符串
    
    # 增加晶圆#作为一列，好几片数据放到一块时，便于区分
    df['waferNum'] = f'{list(sample_ids)[index]}#'        
    # 放到前面
    df = pd.concat([df.iloc[:, -1], df.iloc[:, :-1]], axis=1)
    #df=df[list(df.columns[-1:])+list(df.columns[:-1])]
        
    
    #保存数据
    # 首次写入创建文件，后续追加模式
    mode = 'w' if index==0 else 'a'
    header = (index==0)  # 仅第一次写入列名
    
    # 保存到CSV（注意index=False避免重复索引）
    # df.to_csv(f"{base_filename}_{i+1}.csv", 
    #          mode=mode,
    #          header=header,
    #          index=False)
    df.to_csv( os.path.join(path, all_Pcs_TJparam_detail), mode=mode,
              index=False, header=header)
    print(f"保存完了{list(sample_ids)[index]}#")
        

print(f"处理完成! 结果保存至: {os.path.join(path, all_Pcs_TJparam_detail)}")
