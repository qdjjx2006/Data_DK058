# -*- coding: utf-8 -*-
"""
Created on Tue May 27 11:19:25 2025

@author: pc
"""

import pandas as pd
import numpy as np
import glob
import re
import openpyxl
import os

os.chdir(r'C:\Users\14333\vs_code_DK058data\DK054_5x25')
ResultFile = r'C:\Users\14333\vs_code_DK058data\DK054_5x25\Total_RFB_TJ.xlsx'
# 获取当前文件夹下所有CSV文件
#file_dict_list = glob.glob(r'C:\Users\14333\vs_code_DK058data\DK054_5x25\**/*.csv')
# 法1：读取全部csv文件到列表，os.scandir()判断出文件夹和文件，根据文件夹分类csv，形成不同批次的列表

# 法2：根据os.scandir()的文件夹变量，分别读取不同批次的列表，形成字典
file_dict_list = { }
with os.scandir() as entries:
    for entry in entries:
        if not entry.is_file():
            Xdir_filelist = glob.glob(os.path.join(entry,'*.csv'))
            # file_dict_list_sort = sorted(file_dict_list, key=lambda s: abs(int(s[-6:-4])))
            # 文件名列表根据尾数字排序
            # match = re.search(r'-(\d+).csv$', file_dict_list)
            try:
                file_dict_list[entry.name] = sorted(Xdir_filelist, key=lambda FilePath: int(re.search(r'-(\d+)\.csv$', FilePath).group(1)))
            except:
                None
             

# 该部分为了产生不同尾缀的统计文件（该部分暂未用）
# result_list = glob.glob(r'CSF40_TJ_DATA*.xlsx')
# try:
#     File_num_list = [0] + [re.search(r'CSF40_\w*(\d+)\.xlsx$', FilePath).group(1) for FilePath in result_list]
# except:
#     None
# finally:
#     File_num_list = [int(s) for s in File_num_list]
#     #Wafer_num_list.sort()
#     ResultFile = r"CSF40_TJ_DATA_FT_err" + str(max(File_num_list)+1) + ".xlsx"
  

"""
with open(r'file_dict_list_name_backup.txt', 'r+', encoding="utf-8") as f:
    try:
        read_data = f.read()
    except:
        print("File is empty")
    for str_file in file_dict_list:
        f.write( str_file )
        f.write('\n')
print(f.closed)
        
for str_file in file_dict_list:
    os.rename(str(str_file),str(str_file[:20]+'.csv'))
"""


# 初始化一个空的DataFrame来存储结果
combined_q_data = pd.DataFrame()    
index_order = list(range(16)[::-1])+list(range(16,32)[::-1])+[32,33]    #CSF情况总数 15,14, ..., 1,0,31,30,...,17,16,32,33

combined_TJ = pd.DataFrame(index=index_order)
combined_TJ.index.name = "CSF40"

combined_TJ_Rfb = pd.DataFrame()
combined_TJ_Rfb.index.name = "Rfb"

Wafer_num = 0



# 提取出片号，形成列表————（该部分暂未用）
# try:
#     Wafer_num_list = [int(re.search(r'-(\d+)\.csv$', FilePath).group(1)) for FilePath in file_dict_list]
# except:
#     None
# # 排序
# Wafer_num_list.sort()   # 后面用Wafer_num调用，多此一举



def StrToValue(x):
    a = x.replace('=', '', -1)
    b = a.replace('"', '', -1)
    
    """try:
        c = float(b)
    except ValueError:
        c = float('nan')
        
    return c"""

    d = b.replace('.', '', -1)
    if d.isdigit():
        c = float(b)
    else:
        c = float('nan')
        
    return c

def BinToDec(x):       
    try:
        # c = int(str(x),2)
        c = int(x,2) # as read_csv: dtype={'AA': str}
    except ValueError:
        c = x
        
    return c
 
# 定义一个转换函数，将输入转换为float
def convert_to_float(x):
    try:
        return float(x)
    except ValueError:
        # print("error string:",x,ascii(x))
        return np.nan  # 或者返回其他你认为合适的值，例如0或np.nan
 
# 读取CSV文件，并使用converters参数应用转换函数到所有数值字符串列
# df = pd.read_csv('yourfile.csv', converters={col: convert_to_float for col in df.select_dtypes(include=['object']).columns})

def Adjust_format(filename, ws_name):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    
    # 先用pandas输出基础数据
    # df.to_excel(filename, index=False)
    
    # 用openpyxl处理格式
    wb = load_workbook(filename)
    ws = wb[ws_name]
    
    # 设置自动换行
    wrap_align = Alignment(wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_align
    
    # 按内容调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
        
    # 创建居中对齐样式
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 设置A1单元格
    cell = ws['A1']
    cell.value = "Rfb参数统计"
    cell.font = Font(size=14, bold=True)  # 14号加粗字体
    cell.alignment = center_alignment

    # 遍历A列所有单元格设置样式
    # for cell in ws['A1']:
    #     cell.alignment = center_alignment

    # 遍历所有行和列设置样式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    wb.save(filename)

def pd_read_csv(File):
    df = pd.read_csv(File, header=14, usecols=['AA','Vdd', 'Rfb'], skiprows=0, dtype={'AA': str})
    # names=['A','B','C'], nrows=10,
    # ,'Rfb':float  ##空字符串、文本等会抛出错误
    df = df.iloc[3:]
    #df.loc[:,'AA'][df.loc[:,'AA']==''] = 32
    df.loc[:,'AA'].apply(lambda x: x if x != '' else 32)
    df.loc[:,'AA'].apply(lambda x: x if x != ' ' else 33)
    
    # 索引设置值
    '''df.reset_index(drop=False, inplace = True)
    df['index'] = df['index'] -2;
    df.set_index(['index'],inplace = True)'''
    df.index -= 2   # same as last 3 lines
    
    # 列重命名
    # df.rename(columns={'AA':'CSF40'},inplace = True)
    df.rename({'AA':'CSF40','index':'ChipNum'}, axis=1, inplace = True)
    # df.rename(str.lower, axis='columns',inplace = True)
    
    # 列组命名，索引命名
    df.index.name = 'index_0'
    df.columns.name = 'Pram'
    return df


for Xdir,Xfilelist in file_dict_list.items():
    for File in Xfilelist:
        # 读取CSV文件
        df = pd_read_csv(File)
    
        # 提取Q列数据（从第8行开始）
        #q_data = df['7:Rfbmax'].iloc[7:].reset_index(drop=True).to_frame(name='Q')    
        #q_data = df.iloc[3:,:][PramTJ].dropna().reset_index(drop=True)#.to_frame(name='Rfb')
        #q_data = df.iloc[3:,:].dropna().reset_index(drop=True)#.to_frame(name='Rfb')
        q_data = df['CSF40']
        
        #q_data = q_data.astype(float)
        #q_data = q_data.apply(lambda x: x.replace('=','',-1))
        #q_data = q_data.apply(lambda x: x.replace('"','',-1))
        #q_data = q_data.apply(lambda x: float(x) if x.replace('=', '', 1).isdigit() else float('nan'))
        q_data = q_data.apply(BinToDec).to_frame(name='CSF40参数统计')
        q_data[q_data==' '] = 32
        q_data[q_data==''] = 33
    
    
        TJ = q_data['CSF40参数统计'].value_counts()
        for i in index_order:
            if i not in TJ.index:
                TJ[i] = 0
                
        TJ = TJ.loc[index_order].to_frame(name='CSF40统计' + re.search(r'-(\d+)\.csv$', File).group(1) + '#')
        #TJ = TJ.T
        
        # 将数据添加到总DataFrame    
        # combined_q_data = pd.concat([combined_q_data, q_data], ignore_index=True)
        
        # Sheet2
        # TJ_column_name = 'CSF40统计' + re.search(r'-(\d+)\.csv$', File).group(1) + '#'
        TJ_column_name = re.search(r'-(\d+)\.csv$', File).group(1) + '#'
        combined_TJ[TJ_column_name] = TJ#[TJ_column_name]
        
        # Sheet3
        # df.loc[:,'Rfb'][df['Rfb']==''] = 0
        # df.loc[:,'Rfb'][df['Rfb']==' '] = 0
        # df['Rfb'] = df['Rfb'].astype(float) #空字符串、文本等会抛出错误
            
        # 字符转换为float
        df.loc[:,'Rfb'] = df['Rfb'].apply(convert_to_float)
        
        # df['Rfb'] = pd.to_numeric(df['Rfb'], errors='coerce')  # 将无法转换的值设置为NaN
        
        # float精度设置
        df['Rfb'] = round(df['Rfb'],1)
        '''df['Rfb'] = df['Rfb'].round(2)
        df['Rfb'] = np.around(df['Rfb'], decimals=1)
        
        df['A'] = df['A'].apply(lambda x: format(x, '.2f'))
        df['A'] = df['A'].apply(lambda x: "{:.2f}".format(x))
        
        df.to_csv('output.csv', index=False, float_format='%.2f')   # 保存时转换
        '''
        
        # TJ = pd.cut(df['Rfb'], bins = [10,29] + list(range(30,35)) + [35,40]).value_counts()
        TJ = pd.cut(df['Rfb'], bins = [10,28,29,30] + list(np.arange(30.1, 33, 0.1).round(1)) + [33,34,35,40]).value_counts().sort_index()
        temp = df['Rfb'].sort_values().reset_index(drop = True )
        # TJ = pd.cut(temp[temp.searchsorted(25):temp.searchsorted(34)], bins = 20).value_counts().sort_index()
        # TJ = pd.cut(temp[temp>25][temp<34], bins = 20).value_counts().sort_index()
        # TJ_column_name = {'Rfb':'Rfb参数统计'}
        # TJ = pd.cut(df['Rfb'], bins = 10).value_counts()
        # TJ = pd.qcut(df['Rfb'], q = 10).value_counts()
        # combined_TJ_Rfb['Rfb统计'+str(Wafer_num_list[Wafer_num])+'#'] = TJ
        # combined_TJ_Rfb['Rfb统计' + re.search(r'-(\d+)\.csv$', File).group(1) + '#'] = TJ
        combined_TJ_Rfb.loc[:, TJ_column_name ] = TJ
                
        Wafer_num += 1
    def style_negative(v, props=''):
        return props if v < 0 else None
    with pd.ExcelWriter(
        ResultFile,
        mode="a",
        engine="openpyxl",
        if_sheet_exists="replace") as writer:
        '''combined_q_data.style.\
            map(style_negative, props='color:red;').\
            highlight_max(axis=0).\
            to_excel(writer, index=True, sheet_name="Sheet1", engine='openpyxl')'''
        # combined_TJ.style.\
        #     highlight_max(axis=0).\
        #     to_excel(writer, index=True, sheet_name=Xdir, engine='openpyxl')
            # map(lambda v: 'color:blue;' if (v < 1e6) and (v > 1e3) else None).\
            # .format("{:.2f}")
        combined_TJ_Rfb.style.\
            highlight_max(axis=0).\
            to_excel(writer, index=True, sheet_name=Xdir, engine='openpyxl')
        
    Adjust_format(ResultFile, Xdir)
    
#labels = pd.cut(combined_q_data.Rfb,[.6,1.4,2.2,3,3.8,4.6,4.7,4.8,4.9,5,5.1,5.2,5.3,5.4,5.7,6.2,7])
#grouped = combined_q_data.groupby([labels],observed=False)


#TJ = combined_q_data['CSF40'].value_counts().loc[index_order].to_frame(name='CSF40统计')

#combined_q_data.to_excel('Rfb_data.xlsx', index=True, sheet_name="Sheet1")

# 创建一个Workbook对象，这相当于一个Excel文件
# wb = openpyxl.Workbook()
# 激活当前工作表
# ws = wb.active
 
# 改变工作表的名称
# ws.title = "Sheet1"
# ws.title = re.search(r'-(\d+)\.csv$', result_list[0]).group(1)
 
''' 
# 在单元格A1中写入数据
ws['A1'] = "Hello"
ws['B1'] = "World"
 
# 也可以使用append方法添加行
ws.append([1, 2, 3])'''
 
# 保存文件
# wb.save(ResultFile)'''

print("数据处理完成，结果已保存到",ResultFile)
#df['A'] = df['A'].apply(lambda x: float(x) if x.replace('.', '', 1).isdigit() else float('nan'))
"""
def make_pretty(styler):
    styler.set_caption("Weather Conditions")
    styler.format(rain_condition)
    styler.format_index(lambda v: v.strftime("%A"))
    styler.background_gradient(axis=None, vmin=1, vmax=5, cmap="YlGnBu")
    return styler
"""
